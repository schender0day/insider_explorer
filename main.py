import requests
import json
from datetime import datetime, timedelta
import openpyxl


class InsiderUpdates:
    def __init__(self, api_token: str, symbol: str):
        self.api_token = api_token
        self.symbol = symbol


    def filter_by_value(self, data, value_lower_bound: float):
        filtered_data = []
        us_exchange = ['NYSE', 'NAS']
        for item in data:
            # Only include transactions with type "P" (Purchase)
            if item['type'] != "S" and item['exchange'] in us_exchange:
                # Calculate the value of the transaction
                trans_value = item['cost'] * item['trans_share']
                # Only include transactions with value greater than the lower bound

                if trans_value >= value_lower_bound:
                    filtered_data.append(item)
                trans_value = '{:,.2f}'.format(trans_value)
                # item['total_value'] = trans_value
                item['cost'] = trans_value
        return filtered_data


    def get_data(self, date: str = None):
        if date is None:
            today = datetime.now().date()
            start_date = today
        else:
            start_date = date

        endpoint = f"https://api.gurufocus.com/public/user/{self.api_token}/insider_updates?page=1&date={start_date}"

        # Make the API request
        response = requests.get(endpoint)

        # Check if the request was successful
        if response.status_code == 200:
            # Parse the JSON data from the response
            data = response.json()
            return data


def combine_responses(responses):
    combined_response = []
    for response in responses:
        combined_response += response
    return combined_response

def get_real_time_stock_price(symbol):
    with open("api_token_free.txt", "r") as f:
        api_key = f.read().strip()

    url = f"https://www.alphavantage.co/query?function=GLOBAL_QUOTE&symbol={symbol}&apikey={api_key}"

    response = requests.get(url)
    data = response.json()

    if "Global Quote" in data:
        price = data["Global Quote"]["05. price"]
        return float(price)
    else:
        return None
def get_rank_profitability(api_token, symbol):
    endpoint = f'https://api.gurufocus.com/public/user/{api_token}/stock/{symbol}/summary'
    response = requests.get(endpoint)
    data = response.json()['summary']['general']
    rank_profitability = data['rank_profitability']
    rank_financial_strength = data['rank_profitability']
    rank_gf_value = data['rank_gf_value']
    cur_price = get_real_time_stock_price(symbol)
    rank_momentum = data['rank_momentum']

    return rank_profitability, rank_financial_strength, rank_gf_value, cur_price


def open_link(symbol):
    symbol = symbol.split(":")[1]
    link = f"https://www.google.com/search?q={symbol}"
    link += "+stock"
    return link

def export_to_excel(data):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Symbol", "Exchange", "Position", "Date", "Type", "Trans Share", "Final Share", "Price", "Cost", "Insider", "Link", "Rank Profitability", "Rank Financial Strength", "Rank GF Value", "Current Price", "Increased By (%)"])
    for item in data:
        worksheet.append([item["symbol"], item["exchange"], item["position"], item["date"], item["type"], item["trans_share"], item["final_share"], item["price"], item["cost"], item["insider"], item["link"], item["rank_profitability"], item["rank_financial_strength"], item["rank_gf_value"], item["cur_price"]])
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=12, max_col=12):
        for cell in row:
            cell.hyperlink = cell.value
            cell.value = f"{cell.value}"
    workbook.save("insider_updates.xlsx")



def get_api_token(file_path):
    with open(file_path, "r") as f:
        api_token = f.read().strip()
    return api_token


def get_filtered_data_list(insider_updates):
    date = datetime.now().date()
    filtered_data_list = []
    for i in range(2):
        data = insider_updates.get_data(date.strftime('%Y-%m-%d'))
        filtered_data = insider_updates.filter_by_value(data, 200000)
        filtered_data_list.append(filtered_data)
        date -= timedelta(days=3)
    return filtered_data_list


def process_response(response, api_token):
    processed_response = []
    for item in response:
        item['link'] = open_link(item['symbol'])
        symbol = item['symbol'].split(":")[1]
        rank_profitability = get_rank_profitability(api_token, symbol)
        item['rank_profitability'], item['rank_financial_strength'], item['rank_gf_value'], item['cur_price']= rank_profitability
        # item['increased_by'] = round(100 * (float(item['cur_price']) - float(item['price'])) / float(item['price']), 2)

        if int(item['rank_profitability']) + int(item['rank_financial_strength']) + int(item['rank_gf_value']) >= 23:
            processed_response.append(item)
    return processed_response

def main():
    api_token = get_api_token("api_token.txt")

    insider_updates = InsiderUpdates(api_token, "AAPL")

    filtered_data_list = get_filtered_data_list(insider_updates)

    combined_response = combine_responses(filtered_data_list)
    processed_response = process_response(combined_response, api_token)


    print(json.dumps(processed_response, indent=4))
    export_to_excel(processed_response)
    print("Number of transactions found:", len(processed_response))


if __name__ == '__main__':
    main()
